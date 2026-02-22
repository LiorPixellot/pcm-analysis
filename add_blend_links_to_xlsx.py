#!/usr/bin/env python3
"""
Add clickable blend image hyperlinks to the movement xlsx.

For each row with calibration_movement, locates the corresponding blend image
and adds a file:/// hyperlink in a new 'blend_image' column.

Usage:
    python add_blend_links_to_xlsx.py <data_dir> [--xlsx <path>] [--output <path>]
    python add_blend_links_to_xlsx.py data/16_2_linux_s2
    python add_blend_links_to_xlsx.py data/16_2_linux_s2 --xlsx output_dir/PQS_blur_by_venue_with_movement_16_2_linux_s2.xlsx
"""

import argparse
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font


def find_event_with_movement(venue_path: Path) -> Path | None:
    """Find the event directory that contains movement/."""
    if not venue_path.is_dir():
        return None
    for entry in sorted(venue_path.iterdir()):
        if entry.is_dir() and (entry / "movement").is_dir():
            return entry
    return None


def main():
    parser = argparse.ArgumentParser(description="Add blend image hyperlinks to movement xlsx")
    parser.add_argument("data_dir", type=Path, help="Data directory (e.g. data/16_2_linux_s2)")
    parser.add_argument("--xlsx", type=Path, default=None, help="Input xlsx (default: auto from data_dir name)")
    parser.add_argument("--output", type=Path, default=None, help="Output xlsx (default: overwrite input)")
    args = parser.parse_args()

    data_dir = args.data_dir
    if not data_dir.is_dir():
        print(f"Error: {data_dir} is not a directory")
        sys.exit(1)

    data_dir_name = data_dir.name
    xlsx_path = args.xlsx or (Path("output_dir") / f"PQS_blur_by_venue_with_movement_{data_dir_name}.xlsx")
    output_path = args.output or xlsx_path

    print(f"Loading {xlsx_path}...")
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    venue_col = headers.index("PIXELLOT_VENUE_ID") + 1
    cal_col = headers.index("calibration_movement") + 1
    # Image columns: current pano, reference pano, blend
    image_cols = ["current_image", "reference_image", "blend_image"]
    suffixes = ["current", "reference", "blend"]

    # Add or find each column
    col_indices = {}
    next_col = len(headers) + 1
    for col_name in image_cols:
        if col_name in headers:
            col_indices[col_name] = headers.index(col_name) + 1
        else:
            col_indices[col_name] = next_col
            ws.cell(row=1, column=next_col, value=col_name)
            next_col += 1

    link_font = Font(color="0000FF", underline="single")

    counts = {col: 0 for col in image_cols}
    missing = 0
    no_cal = 0

    # Cache event dirs per venue to avoid repeated filesystem lookups
    event_cache = {}

    for row_idx in range(2, ws.max_row + 1):
        venue_id = ws.cell(row=row_idx, column=venue_col).value
        calibration = ws.cell(row=row_idx, column=cal_col).value

        if not venue_id or not calibration:
            no_cal += 1
            continue

        calibration = str(calibration).lower()

        # Find event dir with movement
        if venue_id not in event_cache:
            event_cache[venue_id] = find_event_with_movement(data_dir / venue_id)

        event_dir = event_cache[venue_id]
        if event_dir is None:
            missing += 1
            continue

        movement_dir = event_dir / "movement"
        row_has_any = False

        for col_name, suffix in zip(image_cols, suffixes):
            fname = f"movement_{venue_id}_pano_{calibration}_{suffix}.jpg"
            fpath = movement_dir / fname

            cell = ws.cell(row=row_idx, column=col_indices[col_name])
            if fpath.exists():
                cell.value = fname
                cell.hyperlink = fpath.resolve().as_uri()
                cell.font = link_font
                counts[col_name] += 1
                row_has_any = True

        if not row_has_any:
            missing += 1

    wb.save(output_path)
    print(f"\nSaved to {output_path}:")
    for col_name in image_cols:
        print(f"  Rows with {col_name}: {counts[col_name]}")
    print(f"  Rows missing all images: {missing}")
    print(f"  Rows without calibration: {no_cal}")


if __name__ == "__main__":
    main()

#!/usr/bin/env python3
"""Add clickable image hyperlink columns to complete_annotation.csv, output as xlsx."""

import csv
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

CSV_PATH = Path("complete_annotation.csv")
DATA_DIR = Path("annotation_data")
OUTPUT_PATH = Path("complete_annotation.xlsx")


def find_focus_dir(date_str, venue_id):
    """Find the focus directory for a venue, picking the first event if multiple exist."""
    venue_dir = DATA_DIR / date_str / venue_id
    if not venue_dir.exists():
        return None
    event_dirs = sorted(d for d in venue_dir.iterdir() if d.is_dir())
    if not event_dirs:
        return None
    focus_dir = event_dirs[0] / "focus"
    if not focus_dir.exists():
        return None
    return focus_dir


def main():
    if not CSV_PATH.exists():
        print(f"Error: {CSV_PATH} not found")
        sys.exit(1)

    with open(CSV_PATH, newline="") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames
        rows = list(reader)

    # Remove old path columns if they exist from previous run
    image_cols = ["joined_image", "cam0_image", "cam1_image"]
    base_fields = [f for f in fieldnames if f not in image_cols]
    out_fieldnames = base_fields + image_cols

    wb = Workbook()
    ws = wb.active
    ws.title = "Annotations"

    # Write header
    for col_idx, name in enumerate(out_fieldnames, 1):
        ws.cell(row=1, column=col_idx, value=name)

    link_font = Font(color="0000FF", underline="single")

    found = 0
    missing = 0

    for row_idx, row in enumerate(rows, 2):
        date_str = row["Checkup date"].split(" ")[0]
        venue_id = row["System ID"].strip()
        focus_dir = find_focus_dir(date_str, venue_id)

        # Write base columns
        for col_idx, field in enumerate(base_fields, 1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(field, ""))

        # Write image hyperlink columns
        if focus_dir:
            cam0_path = focus_dir / "CAM0_1.jpg"
            if not cam0_path.exists():
                cam0_path = focus_dir / "CAM0_1_rot.jpg"
            cam1_path = focus_dir / "CAM1_1.jpg"
            if not cam1_path.exists():
                cam1_path = focus_dir / "CAM1_1_rot.jpg"
            images = {
                "joined_image": focus_dir / "joined_0_1.jpg",
                "cam0_image": cam0_path,
                "cam1_image": cam1_path,
            }
            for col_name in image_cols:
                col_idx = out_fieldnames.index(col_name) + 1
                img_path = images[col_name]
                if img_path.exists():
                    cell = ws.cell(row=row_idx, column=col_idx, value=img_path.name)
                    cell.hyperlink = img_path.resolve().as_uri()
                    cell.font = link_font
            found += 1
        else:
            missing += 1

    wb.save(OUTPUT_PATH)
    print(f"Saved {OUTPUT_PATH}: {found} rows with images, {missing} rows without data")


if __name__ == "__main__":
    main()

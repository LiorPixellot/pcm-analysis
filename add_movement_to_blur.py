#!/usr/bin/env python3
"""
Add movement data from data_11_2 movement.json files to PQS_blur_by_venue.xlsx.

For each venue row in the source xlsx, finds the movement.json in the data dir,
and creates one output row per calibration with movement columns added.

Usage:
    python add_movement_to_blur.py <data_dir>
"""

import json
import os
import sys
from pathlib import Path

import openpyxl

SRC_XLSX = Path(__file__).parent / "PQS_blur_by_venue.xlsx"

NEW_COLS = [
    "calibration_movement",
    "movement_indicator",
    "movement_length_cam0",
    "movement_length_cam1",
    "movement_severity",
]


def classify_movement_severity(indicator, length_cam0, length_cam1):
    """
    Classify movement severity based on indicator and movement lengths.
    - indicator 2 or -1: NA
    - indicator 0: use min(cam0, cam1) to determine Error/Warning/Ok
    """
    if indicator is None:
        return None
    if indicator in (2, -1):
        return "NA"
    if indicator == 0:
        lengths = [l for l in (length_cam0, length_cam1) if l is not None]
        if not lengths:
            return None
        max_length = max(lengths)
        if max_length > 40:
            return "ERROR"
        elif max_length >= 10:
            return "WARNING"
        else:
            return "OK"
    return None


def load_movement_data(data_dir: Path) -> dict:
    """
    Scan data_dir for movement.json files.
    Returns: {venue_id: {calibration_lower: {movement_indicator, length_cam0, length_cam1}}}
    """
    result = {}
    for venue_id in os.listdir(data_dir):
        venue_path = data_dir / venue_id
        if not venue_path.is_dir():
            continue
        events = [e for e in os.listdir(venue_path) if (venue_path / e).is_dir()]
        if not events:
            continue
        mov_json = venue_path / events[0] / "movement" / "movement.json"
        if not mov_json.exists():
            continue
        try:
            with open(mov_json) as f:
                data = json.load(f)
        except Exception as e:
            print(f"  Error reading {mov_json}: {e}")
            continue

        venue_calibrations = {}
        for m in data.get("movements", []):
            calib = m.get("calibration", "").lower()
            cam_id = m.get("camera_id")
            if not calib:
                continue
            if calib not in venue_calibrations:
                venue_calibrations[calib] = {
                    "movement_indicator": m.get("movement_indicator"),
                    "movement_length_cam0": None,
                    "movement_length_cam1": None,
                }
            if cam_id == 0:
                venue_calibrations[calib]["movement_length_cam0"] = m.get("movement_length")
            elif cam_id == 1:
                venue_calibrations[calib]["movement_length_cam1"] = m.get("movement_length")
            # movement_indicator should be same for both cameras, take first seen
            if venue_calibrations[calib]["movement_indicator"] is None:
                venue_calibrations[calib]["movement_indicator"] = m.get("movement_indicator")

        result[venue_id] = venue_calibrations
    return result


def main():
    if len(sys.argv) < 2:
        print(f"Usage: python {sys.argv[0]} <data_dir>")
        sys.exit(1)

    data_dir = Path(sys.argv[1])
    if not data_dir.is_dir():
        print(f"Error: {data_dir} is not a directory")
        sys.exit(1)

    data_dir_name = data_dir.name
    dst_xlsx = Path(__file__).parent / "output_dir" / f"PQS_blur_by_venue_with_movement_{data_dir_name}.xlsx"

    print(f"Loading movement data from {data_dir}...")
    movement_data = load_movement_data(data_dir)
    print(f"  Found movement data for {len(movement_data)} venues")

    print(f"Loading {SRC_XLSX}...")
    wb_src = openpyxl.load_workbook(SRC_XLSX, read_only=True)
    ws_src = wb_src.active
    src_headers = [cell.value for cell in ws_src[1]]
    venue_col_idx = src_headers.index("PIXELLOT_VENUE_ID")
    num_src_cols = len(src_headers)

    wb_dst = openpyxl.Workbook()
    ws_dst = wb_dst.active

    # Write headers
    all_headers = src_headers + NEW_COLS
    for c, name in enumerate(all_headers, 1):
        ws_dst.cell(row=1, column=c, value=name)

    out_row = 2
    matched_venues = 0
    unmatched_venues = 0
    total_out_rows = 0

    for src_row in ws_src.iter_rows(min_row=2, values_only=True):
        venue_id = src_row[venue_col_idx]
        if not venue_id:
            continue

        # Copy source columns
        src_values = list(src_row[:num_src_cols])

        venue_movements = movement_data.get(venue_id)
        if not venue_movements:
            # No movement data — write one row with empty movement cols
            for c, val in enumerate(src_values, 1):
                ws_dst.cell(row=out_row, column=c, value=val)
            out_row += 1
            total_out_rows += 1
            unmatched_venues += 1
            continue

        matched_venues += 1
        # One row per calibration
        for calib, mov in venue_movements.items():
            for c, val in enumerate(src_values, 1):
                ws_dst.cell(row=out_row, column=c, value=val)
            # New columns
            base = num_src_cols + 1
            ws_dst.cell(row=out_row, column=base, value=calib)
            ws_dst.cell(row=out_row, column=base + 1, value=mov["movement_indicator"])
            ws_dst.cell(row=out_row, column=base + 2, value=mov["movement_length_cam0"])
            ws_dst.cell(row=out_row, column=base + 3, value=mov["movement_length_cam1"])
            ws_dst.cell(row=out_row, column=base + 4, value=classify_movement_severity(
                mov["movement_indicator"], mov["movement_length_cam0"], mov["movement_length_cam1"]
            ))
            out_row += 1
            total_out_rows += 1

    wb_src.close()
    wb_dst.save(dst_xlsx)

    print(f"\nResults:")
    print(f"  Venues with movement data: {matched_venues}")
    print(f"  Venues without movement:   {unmatched_venues}")
    print(f"  Output rows:               {total_out_rows}")
    print(f"  Saved to: {dst_xlsx}")


if __name__ == "__main__":
    main()

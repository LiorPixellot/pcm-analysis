#!/usr/bin/env python3
"""
Join laplacian threshold CSV with PQS blur-by-venue CSV on venue ID.

Usage:
    python concat_blur.py <laplacian_th.csv> <blur_by_venue.csv> [-o output.xlsx]
"""

import argparse
import csv
import os
import re
import sys

import openpyxl


HYPERLINK_RE = re.compile(r'^=HYPERLINK\("(.+?)"\)$')


def main():
    parser = argparse.ArgumentParser(
        description='Join laplacian threshold CSV with blur-by-venue CSV on venue ID'
    )
    parser.add_argument('laplacian_csv', help='Laplacian threshold calculations CSV (has venue_id column)')
    parser.add_argument('blur_csv', help='Blur by venue CSV or Excel file (has PIXELLOT_VENUE_ID column)')
    parser.add_argument('-o', '--output', default='laplacian_th_with_blur.xlsx',
                        help='Output Excel file (default: same directory as laplacian_csv)')
    args = parser.parse_args()

    if args.output == 'laplacian_th_with_blur.xlsx':
        args.output = os.path.join(os.path.dirname(args.laplacian_csv), 'laplacian_th_with_blur.xlsx')

    # Load blur lookup: PIXELLOT_VENUE_ID -> row
    blur_lookup = {}
    blur_fields = []
    try:
        if args.blur_csv.endswith(('.xlsx', '.xls')):
            wb = openpyxl.load_workbook(args.blur_csv, read_only=True)
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            headers = [str(c) if c is not None else '' for c in next(rows_iter)]
            blur_fields = [col for col in headers if col != 'PIXELLOT_VENUE_ID']
            vid_idx = headers.index('PIXELLOT_VENUE_ID')
            for row in rows_iter:
                vid = str(row[vid_idx]) if row[vid_idx] is not None else ''
                if vid:
                    blur_lookup[vid] = {h: str(row[i]) if row[i] is not None else '' for i, h in enumerate(headers) if h != 'PIXELLOT_VENUE_ID'}
            wb.close()
        else:
            with open(args.blur_csv, 'r', newline='') as f:
                reader = csv.DictReader(f)
                blur_fields = [col for col in reader.fieldnames if col != 'PIXELLOT_VENUE_ID']
                for row in reader:
                    vid = row['PIXELLOT_VENUE_ID']
                    if vid:
                        blur_lookup[vid] = {k: row[k] for k in blur_fields}
    except FileNotFoundError:
        print(f'Error: {args.blur_csv} not found', file=sys.stderr)
        sys.exit(1)
    except (KeyError, ValueError):
        print(f'Error: {args.blur_csv} missing PIXELLOT_VENUE_ID column', file=sys.stderr)
        sys.exit(1)

    print(f'Loaded {len(blur_lookup)} venues from {args.blur_csv}')

    # Join and write output
    matched = 0
    unmatched = 0

    try:
        with open(args.laplacian_csv, 'r', newline='') as infile:
            reader = csv.DictReader(infile)
            out_fields = reader.fieldnames + blur_fields

            wb_out = openpyxl.Workbook()
            ws_out = wb_out.active
            ws_out.append(out_fields)

            for row in reader:
                vid = row['venue_id']
                if vid in blur_lookup:
                    row.update(blur_lookup[vid])
                    matched += 1
                else:
                    for col in blur_fields:
                        row[col] = ''
                    unmatched += 1

                ws_out.append([row.get(f, '') for f in out_fields])

            # Convert =HYPERLINK(...) formulas to real Excel hyperlinks
            if 'joined_image' in out_fields:
                col_idx = out_fields.index('joined_image') + 1  # openpyxl is 1-indexed
                for row_cells in ws_out.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                    cell = row_cells[0]
                    if cell.value:
                        m = HYPERLINK_RE.match(str(cell.value))
                        if m:
                            url = m.group(1)
                            cell.hyperlink = url
                            cell.value = os.path.basename(url)
                            cell.style = 'Hyperlink'

            wb_out.save(args.output)

    except FileNotFoundError:
        print(f'Error: {args.laplacian_csv} not found', file=sys.stderr)
        sys.exit(1)
    except KeyError:
        print(f'Error: {args.laplacian_csv} missing venue_id column', file=sys.stderr)
        sys.exit(1)

    print(f'Matched: {matched}')
    print(f'Unmatched (no blur data): {unmatched}')
    print(f'Output written to: {os.path.abspath(args.output)}')


if __name__ == '__main__':
    main()

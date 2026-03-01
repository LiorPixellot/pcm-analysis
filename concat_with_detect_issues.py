#!/usr/bin/env python3
"""
Join pipeline Excel with detect_issues CSV on venue_id + event_id.

Usage:
    python concat_with_detect_issues.py <pipeline_xlsx> <detect_issues_csv> [-o output.xlsx]
"""

import argparse
import csv
import os
import sys

import openpyxl


def main():
    parser = argparse.ArgumentParser(
        description='Join pipeline Excel with detect_issues CSV on venue_id + event_id'
    )
    parser.add_argument('pipeline_xlsx', help='Pipeline Excel (from concat_with_is_measurable.py)')
    parser.add_argument('detect_issues_csv', help='Detect issues CSV (from detect_issues.py)')
    parser.add_argument('-o', '--output', default='laplacian_th_with_blur_measurable_issues.xlsx',
                        help='Output Excel file (default: same directory as pipeline_xlsx)')
    args = parser.parse_args()

    if args.output == 'laplacian_th_with_blur_measurable_issues.xlsx':
        args.output = os.path.join(os.path.dirname(args.pipeline_xlsx),
                                   'laplacian_th_with_blur_measurable_issues.xlsx')

    # Load detect_issues lookup: (venue_id, event_id) -> row
    issues_lookup = {}
    issues_fields = []
    try:
        with open(args.detect_issues_csv, 'r', newline='') as f:
            reader = csv.DictReader(f)
            all_fields = reader.fieldnames
            # Only add columns not used as join keys
            issues_fields = [c for c in all_fields if c not in ('venue_id', 'event_id')]
            for row in reader:
                vid = row['venue_id']
                eid = row['event_id']
                if vid and eid:
                    issues_lookup[(vid, eid)] = row
    except FileNotFoundError:
        print(f'Error: {args.detect_issues_csv} not found', file=sys.stderr)
        sys.exit(1)
    except KeyError as e:
        print(f'Error: {args.detect_issues_csv} missing column: {e}', file=sys.stderr)
        sys.exit(1)

    print(f'Loaded {len(issues_lookup)} entries from {args.detect_issues_csv}')

    # Load pipeline xlsx
    try:
        wb_in = openpyxl.load_workbook(args.pipeline_xlsx)
    except FileNotFoundError:
        print(f'Error: {args.pipeline_xlsx} not found', file=sys.stderr)
        sys.exit(1)

    ws_in = wb_in.active
    headers = [cell.value for cell in ws_in[1]]

    # Only add columns not already present
    new_fields = [c for c in issues_fields if c not in headers]
    print(f'Columns to add: {new_fields}')

    # Find venue_id and event_id column indices
    try:
        vid_col = headers.index('venue_id')
    except ValueError:
        print(f'Error: {args.pipeline_xlsx} missing venue_id column', file=sys.stderr)
        sys.exit(1)

    try:
        eid_col = headers.index('event_id')
    except ValueError:
        print(f'Error: {args.pipeline_xlsx} missing event_id column', file=sys.stderr)
        sys.exit(1)

    # Append new column headers
    base_cols = len(headers)
    for i, field in enumerate(new_fields):
        ws_in.cell(row=1, column=base_cols + 1 + i, value=field)

    # Join data
    matched = 0
    unmatched = 0

    for row_idx in range(2, ws_in.max_row + 1):
        vid = ws_in.cell(row=row_idx, column=vid_col + 1).value
        vid = str(vid) if vid is not None else ''
        eid = ws_in.cell(row=row_idx, column=eid_col + 1).value
        eid = str(eid) if eid is not None else ''

        key = (vid, eid)
        if key in issues_lookup:
            for i, col in enumerate(new_fields):
                ws_in.cell(row=row_idx, column=base_cols + 1 + i,
                           value=issues_lookup[key].get(col, ''))
            matched += 1
        else:
            for i in range(len(new_fields)):
                ws_in.cell(row=row_idx, column=base_cols + 1 + i, value='')
            unmatched += 1

    wb_in.save(args.output)

    print(f'Matched: {matched}')
    print(f'Unmatched (no detect_issues data): {unmatched}')
    print(f'Output written to: {os.path.abspath(args.output)}')


if __name__ == '__main__':
    main()

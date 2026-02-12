#!/usr/bin/env python3
"""
Join laplacian threshold + blur Excel with measurable CSV on venue_id.

Usage:
    python concat_with_is_measurable.py <blur_xlsx> <measurable_csv> [-o output.xlsx]
"""

import argparse
import csv
import os
import sys

import openpyxl
from openpyxl.utils import get_column_letter


def main():
    parser = argparse.ArgumentParser(
        description='Join blur Excel with measurable CSV on venue_id'
    )
    parser.add_argument('blur_xlsx', help='Laplacian threshold + blur Excel (from concat_blur.py)')
    parser.add_argument('measurable_csv', help='Measurable CSV (from is_measurable.py or run_is_measurable.py)')
    parser.add_argument('-o', '--output', default='laplacian_th_with_blur_and_measurable.xlsx',
                        help='Output Excel file (default: same directory as blur_xlsx)')
    args = parser.parse_args()

    if args.output == 'laplacian_th_with_blur_and_measurable.xlsx':
        args.output = os.path.join(os.path.dirname(args.blur_xlsx), 'laplacian_th_with_blur_and_measurable.xlsx')

    # Load measurable lookup: venue_id -> row
    measurable_lookup = {}
    measurable_fields = []
    try:
        with open(args.measurable_csv, 'r', newline='') as f:
            reader = csv.DictReader(f)
            all_fields = reader.fieldnames
            for row in reader:
                vid = row['venue_id']
                if vid:
                    measurable_lookup[vid] = row
    except FileNotFoundError:
        print(f'Error: {args.measurable_csv} not found', file=sys.stderr)
        sys.exit(1)
    except KeyError:
        print(f'Error: {args.measurable_csv} missing venue_id column', file=sys.stderr)
        sys.exit(1)

    print(f'Loaded {len(measurable_lookup)} entries from {args.measurable_csv}')

    # Load blur xlsx
    try:
        wb_in = openpyxl.load_workbook(args.blur_xlsx)
    except FileNotFoundError:
        print(f'Error: {args.blur_xlsx} not found', file=sys.stderr)
        sys.exit(1)

    ws_in = wb_in.active
    blur_headers = [cell.value for cell in ws_in[1]]

    # Only add columns not already present in blur xlsx
    measurable_fields = [c for c in all_fields if c not in blur_headers]
    print(f'Columns to add: {measurable_fields}')

    # Find venue_id column index
    try:
        vid_col = blur_headers.index('venue_id')
    except ValueError:
        print(f'Error: {args.blur_xlsx} missing venue_id column', file=sys.stderr)
        sys.exit(1)

    # Append new column headers
    base_cols = len(blur_headers)
    for i, field in enumerate(measurable_fields):
        ws_in.cell(row=1, column=base_cols + 1 + i, value=field)

    # Join data
    matched = 0
    unmatched = 0

    for row_idx in range(2, ws_in.max_row + 1):
        vid = ws_in.cell(row=row_idx, column=vid_col + 1).value
        vid = str(vid) if vid is not None else ''
        if vid in measurable_lookup:
            for i, col in enumerate(measurable_fields):
                ws_in.cell(row=row_idx, column=base_cols + 1 + i,
                           value=measurable_lookup[vid].get(col, ''))
            matched += 1
        else:
            for i in range(len(measurable_fields)):
                ws_in.cell(row=row_idx, column=base_cols + 1 + i, value='')
            unmatched += 1

    wb_in.save(args.output)

    print(f'Matched: {matched}')
    print(f'Unmatched (no measurable data): {unmatched}')
    print(f'Output written to: {os.path.abspath(args.output)}')


if __name__ == '__main__':
    main()

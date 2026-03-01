"""
Consolidate two pipeline run XLSXs into one, picking the best row per venue.

Selection rules per venue:
1. Venue only in one run -> use that row
2. Both runs exist, one is_measurable=Yes the other not -> use measurable row
3. Both is_measurable=Yes -> pick less problematic Focus_severity (Ok > Warning > Error > NaN); tie -> run1
4. Neither measurable -> use run1

Output includes source_data_dir column and clickable image hyperlinks.
"""

from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from PIL import Image

RUN1_PATH = "output_dir/2026-02-22_16-09/concat_data/laplacian_th_with_blur_measurable_issues.xlsx"
RUN2_PATH = "output_dir/2026-02-22_14-40/concat_data/laplacian_th_with_blur_measurable_issues.xlsx"
OUTPUT_PATH = "output_dir/consolidated_best_per_venue_16_18_linux_s2.xlsx"

RUN1_DATA_DIR = "data/16_2_linux_s2"
RUN2_DATA_DIR = "data/18_2_linux_s2"

# Lower = better
SEVERITY_RANK = {"Ok": 0, "Warning": 1, "Error": 2}
# NaN gets rank 3 (worst)
DEFAULT_SEVERITY_RANK = 3


def generate_joined(focus_dir: Path) -> Optional[Path]:
    """Generate joined_0_1.jpg from CAM0_1.jpg + CAM1_1.jpg in focus_dir.

    Left half = CAM1 right half, Right half = CAM0 left half, each resized to 800x800.
    Returns path to generated file, or None if CAMs missing.
    """
    joined_path = focus_dir / "joined_0_1.jpg"
    if joined_path.exists():
        return joined_path

    cam0_path = focus_dir / "CAM0_1.jpg"
    cam1_path = focus_dir / "CAM1_1.jpg"
    if not cam0_path.exists() or not cam1_path.exists():
        return None

    cam0 = Image.open(cam0_path)
    cam1 = Image.open(cam1_path)

    w0, h0 = cam0.size
    w1, h1 = cam1.size

    # Right half of CAM1 (inner half facing CAM0)
    cam1_crop = cam1.crop((w1 // 2, 0, w1, h1)).resize((800, 800))
    # Left half of CAM0 (inner half facing CAM1)
    cam0_crop = cam0.crop((0, 0, w0 // 2, h0)).resize((800, 800))

    joined = Image.new("RGB", (1600, 800))
    joined.paste(cam1_crop, (0, 0))
    joined.paste(cam0_crop, (800, 0))
    joined.save(joined_path, quality=85)
    return joined_path


def is_measurable_yes(val):
    return val == "Yes"


def severity_rank(val):
    return SEVERITY_RANK.get(val, DEFAULT_SEVERITY_RANK)


def main():
    df1 = pd.read_excel(RUN1_PATH)
    df2 = pd.read_excel(RUN2_PATH)

    # Deduplicate each by venue_id (keep first)
    df1 = df1.drop_duplicates(subset="venue_id", keep="first")
    df2 = df2.drop_duplicates(subset="venue_id", keep="first")

    print(f"Run1: {len(df1)} unique venues")
    print(f"Run2: {len(df2)} unique venues")

    # Index by venue_id for fast lookup
    df1_idx = df1.set_index("venue_id")
    df2_idx = df2.set_index("venue_id")

    all_venues = df1_idx.index.union(df2_idx.index)
    print(f"Total unique venues: {len(all_venues)}")

    chosen_rows = []
    chosen_sources = []
    stats = {"only_run1": 0, "only_run2": 0, "measurable_wins": 0, "severity_wins": 0, "tie_run1": 0, "neither_run1": 0}

    for venue_id in all_venues:
        in1 = venue_id in df1_idx.index
        in2 = venue_id in df2_idx.index

        if in1 and not in2:
            chosen_rows.append(df1_idx.loc[venue_id])
            chosen_sources.append(RUN1_DATA_DIR)
            stats["only_run1"] += 1
        elif in2 and not in1:
            chosen_rows.append(df2_idx.loc[venue_id])
            chosen_sources.append(RUN2_DATA_DIR)
            stats["only_run2"] += 1
        else:
            row1 = df1_idx.loc[venue_id]
            row2 = df2_idx.loc[venue_id]
            m1 = is_measurable_yes(row1["is_measurable"])
            m2 = is_measurable_yes(row2["is_measurable"])

            if m1 and not m2:
                chosen_rows.append(row1)
                chosen_sources.append(RUN1_DATA_DIR)
                stats["measurable_wins"] += 1
            elif m2 and not m1:
                chosen_rows.append(row2)
                chosen_sources.append(RUN2_DATA_DIR)
                stats["measurable_wins"] += 1
            elif m1 and m2:
                s1 = severity_rank(row1["Focus_severity"])
                s2 = severity_rank(row2["Focus_severity"])
                if s1 <= s2:
                    chosen_rows.append(row1)
                    chosen_sources.append(RUN1_DATA_DIR)
                    stats["tie_run1" if s1 == s2 else "severity_wins"] += 1
                else:
                    chosen_rows.append(row2)
                    chosen_sources.append(RUN2_DATA_DIR)
                    stats["severity_wins"] += 1
            else:
                # Neither measurable
                chosen_rows.append(row1)
                chosen_sources.append(RUN1_DATA_DIR)
                stats["neither_run1"] += 1

    result = pd.DataFrame(chosen_rows)
    result.index.name = "venue_id"
    result = result.reset_index()
    result["source_data_dir"] = chosen_sources
    result.to_excel(OUTPUT_PATH, index=False)

    # Add image hyperlinks using openpyxl
    wb = load_workbook(OUTPUT_PATH)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    vid_col = headers.index("venue_id") + 1
    eid_col = headers.index("event_id") + 1
    src_col = headers.index("source_data_dir") + 1

    # Add hyperlink columns
    image_cols = ["cam0_image", "cam1_image", "joined_image"]
    for i, col_name in enumerate(image_cols):
        ws.cell(row=1, column=len(headers) + 1 + i, value=col_name)

    link_font = Font(color="0000FF", underline="single")
    found = 0
    missing = 0
    joined_generated = 0

    for row_idx in range(2, ws.max_row + 1):
        venue_id = str(ws.cell(row=row_idx, column=vid_col).value or "")
        event_id = str(ws.cell(row=row_idx, column=eid_col).value or "")
        data_dir = str(ws.cell(row=row_idx, column=src_col).value or "")

        if not venue_id or not event_id or not data_dir:
            missing += 1
            continue

        focus_dir = Path(data_dir) / venue_id / event_id / "focus"
        if not focus_dir.exists():
            missing += 1
            continue

        # Generate joined image if it doesn't exist
        joined_path = generate_joined(focus_dir)
        if joined_path:
            joined_generated += 1

        images = {"cam0_image": "CAM0_1.jpg", "cam1_image": "CAM1_1.jpg", "joined_image": "joined_0_1.jpg"}
        for i, col_name in enumerate(image_cols):
            img_path = focus_dir / images[col_name]
            if img_path.exists():
                cell = ws.cell(row=row_idx, column=len(headers) + 1 + i, value=img_path.name)
                cell.hyperlink = img_path.resolve().as_uri()
                cell.font = link_font
        found += 1

    wb.save(OUTPUT_PATH)
    print(f"Image hyperlinks: {found} rows with images, {missing} rows without")
    print(f"Joined images generated: {joined_generated}")

    print(f"\nWrote {len(result)} rows to {OUTPUT_PATH}")
    print(f"\nSelection summary:")
    print(f"  Only in run1:                  {stats['only_run1']}")
    print(f"  Only in run2:                  {stats['only_run2']}")
    print(f"  Measurable wins over non-meas: {stats['measurable_wins']}")
    print(f"  Better severity wins:          {stats['severity_wins']}")
    print(f"  Equal severity (run1 default): {stats['tie_run1']}")
    print(f"  Neither measurable (run1):     {stats['neither_run1']}")


if __name__ == "__main__":
    main()

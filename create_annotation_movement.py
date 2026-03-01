#!/usr/bin/env python3
"""
Create complete_annotation_movement.xlsx from Annotators - Working Table 2025.xlsx.

Extracts movement-related columns from sheets POC_1.0, POC_1.1, POC_1.2,
normalizes column names, preserves Movement image hyperlinks, and writes
a single combined xlsx.

Usage:
    python create_annotation_movement.py
"""

from pathlib import Path

from openpyxl import Workbook, load_workbook

SRC_XLSX = Path(__file__).parent / "Annotators - Working Table 2025.xlsx"
DST_XLSX = Path(__file__).parent / "complete_annotation_movement.xlsx"

SHEETS = ["POC_1.0", "POC_1.1", "POC_1.2"]

# Output columns in order
OUT_COLS = [
    "Checkup date",
    "System ID",
    "Calibration",
    "Setup quality",
    "Indoor / Outdoor",
    "Far / Close",
    "source_sheet",
    "Movement image",
    "Movement severity",
    "Is Measurable Movement Image (Yes/No)",
    "Movement issues (ERROR/OK/WARNING)",
    "Pixels_distance",
    "Movement length (pixels)",
]

# Per-sheet mapping: output column name -> source column name
# (only entries that differ from the output name)
COL_ALIASES = {
    "POC_1.0": {
        "Movement severity": "Movement severity",
        "Is Measurable Movement Image (Yes/No)": "Is Measurable \nMov(Yes/No)ement Image",
        "Movement issues (ERROR/OK/WARNING)": "Movement issues\n(ERROR/OK/WARNING)",
    },
    "POC_1.1": {
        "Setup quality": "Setup severity",
        "Movement severity": "Movement severity",
        "Is Measurable Movement Image (Yes/No)": "Is Measurable \nMovement Image\n(Yes/No)",
        "Movement issues (ERROR/OK/WARNING)": "Movement issues\n(ERROR/OK/WARNING)",
    },
    "POC_1.2": {
        "Movement severity": "Movement severity",
        "Is Measurable Movement Image (Yes/No)": "Is Measurable \nMovement Image\n(Yes/No)",
        "Movement issues (ERROR/OK/WARNING)": "Movement issues\n(ERROR/OK/WARNING)",
    },
}


def build_col_index(header: list[str | None]) -> dict[str, int]:
    """Map column name -> 0-based index, skipping None headers."""
    return {name: i for i, name in enumerate(header) if name is not None}


def main():
    # Need non-read-only mode to access hyperlinks
    print(f"Loading {SRC_XLSX}...")
    wb_src = load_workbook(SRC_XLSX)

    wb_dst = Workbook()
    ws_dst = wb_dst.active
    ws_dst.title = "movement_annotations"

    # Write header
    for c, name in enumerate(OUT_COLS, 1):
        ws_dst.cell(row=1, column=c, value=name)

    out_row = 2
    total_per_sheet = {}

    for sheet_name in SHEETS:
        ws = wb_src[sheet_name]
        header = [cell.value for cell in ws[1]]
        col_idx = build_col_index(header)

        aliases = COL_ALIASES.get(sheet_name, {})
        sheet_count = 0

        for src_row in ws.iter_rows(min_row=2):
            # Skip empty rows (no System ID)
            sys_id_cell = src_row[col_idx["System ID"]]
            if not sys_id_cell.value:
                continue

            for c, out_col in enumerate(OUT_COLS, 1):
                if out_col == "source_sheet":
                    ws_dst.cell(row=out_row, column=c, value=sheet_name)
                    continue

                # Resolve source column name via alias or direct match
                src_col_name = aliases.get(out_col, out_col)
                if src_col_name not in col_idx:
                    continue

                src_cell = src_row[col_idx[src_col_name]]
                dst_cell = ws_dst.cell(row=out_row, column=c, value=src_cell.value)

                # Preserve hyperlinks (Movement image has S3 URLs)
                if src_cell.hyperlink:
                    dst_cell.hyperlink = src_cell.hyperlink.target
                    dst_cell.style = "Hyperlink"

            out_row += 1
            sheet_count += 1

        total_per_sheet[sheet_name] = sheet_count
        print(f"  {sheet_name}: {sheet_count} rows")

    wb_src.close()
    wb_dst.save(DST_XLSX)

    total = sum(total_per_sheet.values())
    print(f"\nTotal: {total} rows")
    print(f"Saved to: {DST_XLSX}")


if __name__ == "__main__":
    main()

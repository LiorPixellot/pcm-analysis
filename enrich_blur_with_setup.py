#!/usr/bin/env python3
"""
Enrich PQS_blur_by_venue.xlsx with geometric setup data from setup.json files.

For each venue, finds the first event with a setup/setup.json, locates the
entry with max_field_area="MAX", and extracts setup_sport_type,
max_camera_overlap, and all_spares.

Usage:
    python enrich_blur_with_setup.py --dataset data_11_2
    python enrich_blur_with_setup.py --dataset data_11_2 --blur-xlsx PQS_blur_by_venue.xlsx
"""

import json
import argparse
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Font


def read_blur_xlsx(xlsx_path: Path) -> tuple[list[str], list[dict]]:
    """Read PQS_blur_by_venue.xlsx and return (headers, list of row dicts)."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        wb.close()
        return [], []

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    data = []
    for row in rows[1:]:
        row_dict = {}
        for i, val in enumerate(row):
            if i < len(headers):
                row_dict[headers[i]] = val if val is not None else ""
        data.append(row_dict)

    wb.close()
    return headers, data


def find_venue_setup_json(dataset_dir: Path, venue_id: str) -> Optional[Path]:
    """Find the first event with a setup/setup.json for the given venue."""
    venue_dir = dataset_dir / str(venue_id)
    if not venue_dir.is_dir():
        return None

    for event_path in sorted(venue_dir.iterdir()):
        if not event_path.is_dir():
            continue
        setup_json = event_path / "setup" / "setup.json"
        if setup_json.is_file():
            return setup_json

    return None


def extract_max_setup(setup_json_path: Path) -> tuple[Optional[dict], bool, str]:
    """Parse setup.json and return (entry, used_fallback, setup_joined_image).

    First tries the entry with max_field_area='MAX'. If none found, falls back
    to the entry with the largest field_area. Returns (None, False, '') only if the
    setups array is empty.
    """
    with open(setup_json_path) as f:
        data = json.load(f)

    setup_joined_image = data.get("setup_joined_image", "")

    setups = data.get("setups", [])
    if not setups:
        return None, False, setup_joined_image

    for entry in setups:
        if entry.get("max_field_area") == "MAX":
            return entry, False, setup_joined_image

    # Fallback: pick entry with the largest field_area
    best = max(setups, key=lambda e: e.get("field_area", 0))
    return best, True, setup_joined_image


def compute_can_improve_by_zoom(left_spare, right_spare, max_camera_overlap) -> str:
    """Return 'Yes' if avg spare / overlap ratio is in [0.7, 1.3], else 'No'. Empty on bad data."""
    try:
        ls = float(left_spare)
        rs = float(right_spare)
        mco = float(max_camera_overlap)
    except (ValueError, TypeError):
        return ""
    if mco == 0:
        return ""
    ratio = (ls + rs) / mco / 2
    return "Yes" if 0.7 <= ratio <= 1.3 else "No"


def compute_decision(setup_severity: str, can_improve_by_zoom: str) -> str:
    """Derive action decision from setup severity and zoom improvability."""
    if setup_severity == "Error":
        return "maintain_remote" if can_improve_by_zoom == "Yes" else "maintain_on_site"
    if setup_severity == "Warning":
        return "watch"
    if setup_severity == "Ok":
        return "ok"
    return "NA"


def compute_setup_severity(all_spares) -> str:
    if all_spares == "" or all_spares is None:
        return "Aborted"
    try:
        value = float(all_spares)
    except (ValueError, TypeError):
        return "Aborted"
    if value > 3000:
        return "Error"
    if value >= 2000:
        return "Warning"
    return "Ok"


def main():
    parser = argparse.ArgumentParser(
        description="Enrich PQS_blur_by_venue.xlsx with setup.json geometry data"
    )
    parser.add_argument("--dataset", default="data_11_2", help="Data directory (default: data_11_2)")
    parser.add_argument("--blur-xlsx", default="PQS_blur_by_venue.xlsx", help="Input blur xlsx")
    parser.add_argument("-o", "--output", default="output_dir/pqs_blur_by_venue_with_setup.xlsx",
                        help="Output xlsx path")
    args = parser.parse_args()

    dataset_dir = Path(args.dataset)
    xlsx_path = Path(args.blur_xlsx)
    output_path = Path(args.output)

    if not xlsx_path.exists():
        print(f"Error: {xlsx_path} not found")
        return

    if not dataset_dir.is_dir():
        print(f"Error: dataset directory {dataset_dir} not found")
        return

    headers, rows = read_blur_xlsx(xlsx_path)
    if not rows:
        print("No data rows found in xlsx")
        return

    new_columns = [
        "setup_sport_type", "max_camera_overlap", "all_spares", "setup_severity",
        "left_spare", "right_spare", "s2_mode_calc", "X_of_center", "Y_from_line",
        "height", "focals_diff", "cam_angle_diff", "can_improve_by_zoom", "decision",
        "setup_join_image",
    ]
    print(f"Read {len(rows)} rows from {xlsx_path}")
    print(f"Looking up setup.json in {dataset_dir}/")

    matched_max = 0
    matched_fallback = 0
    no_venue_dir = 0
    no_setup_json = 0
    no_setups = 0
    empty_venue_id = 0
    pqs_venue_ids = set()

    for row in rows:
        venue_id = str(row.get("PIXELLOT_VENUE_ID", "")).strip()
        if not venue_id:
            for col in new_columns:
                row[col] = ""
            row["setup_severity"] = "Aborted"
            row["decision"] = "NA"
            empty_venue_id += 1
            continue

        pqs_venue_ids.add(venue_id)
        setup_json_path = find_venue_setup_json(dataset_dir, venue_id)
        if setup_json_path is None:
            for col in new_columns:
                row[col] = ""
            row["setup_severity"] = "Aborted"
            row["decision"] = "NA"
            venue_dir = dataset_dir / venue_id
            if not venue_dir.is_dir():
                no_venue_dir += 1
            else:
                no_setup_json += 1
            continue

        max_entry, used_fallback, setup_joined_image = extract_max_setup(setup_json_path)
        if max_entry is None:
            for col in new_columns:
                row[col] = ""
            row["setup_severity"] = "Aborted"
            row["decision"] = "NA"
            row["setup_join_image"] = setup_joined_image
            no_setups += 1
            continue

        row["setup_sport_type"] = max_entry.get("sport_type", "")
        row["max_camera_overlap"] = max_entry.get("max_camera_overlap", "")
        row["all_spares"] = max_entry.get("all_spares", "")
        row["setup_severity"] = compute_setup_severity(row["all_spares"])
        row["left_spare"] = max_entry.get("left_spare", "")
        row["right_spare"] = max_entry.get("right_spare", "")
        row["s2_mode_calc"] = max_entry.get("s2_mode_calc", "")
        row["X_of_center"] = max_entry.get("X_of_center", "")
        row["Y_from_line"] = max_entry.get("Y_from_line", "")
        row["height"] = max_entry.get("height", "")
        row["focals_diff"] = max_entry.get("focals_diff", "")
        row["cam_angle_diff"] = max_entry.get("cam_angle_diff", "")
        row["can_improve_by_zoom"] = compute_can_improve_by_zoom(
            row["left_spare"], row["right_spare"], row["max_camera_overlap"]
        )
        row["decision"] = compute_decision(row["setup_severity"], row["can_improve_by_zoom"])
        row["setup_join_image"] = setup_joined_image
        if used_fallback:
            matched_fallback += 1
        else:
            matched_max += 1

    # Second pass: venues with setup.json in dataset but not in PQS xlsx
    dataset_only = 0
    dataset_only_no_setups = 0
    for venue_dir in sorted(dataset_dir.iterdir()):
        if not venue_dir.is_dir():
            continue
        venue_id = venue_dir.name
        if venue_id in pqs_venue_ids:
            continue
        setup_json_path = find_venue_setup_json(dataset_dir, venue_id)
        if setup_json_path is None:
            continue
        max_entry, used_fallback, setup_joined_image = extract_max_setup(setup_json_path)
        if max_entry is None:
            dataset_only_no_setups += 1
            continue
        dataset_only += 1
        row = {h: "" for h in headers}
        row["PIXELLOT_VENUE_ID"] = venue_id
        row["setup_sport_type"] = max_entry.get("sport_type", "")
        row["max_camera_overlap"] = max_entry.get("max_camera_overlap", "")
        row["all_spares"] = max_entry.get("all_spares", "")
        row["setup_severity"] = compute_setup_severity(row["all_spares"])
        row["left_spare"] = max_entry.get("left_spare", "")
        row["right_spare"] = max_entry.get("right_spare", "")
        row["s2_mode_calc"] = max_entry.get("s2_mode_calc", "")
        row["X_of_center"] = max_entry.get("X_of_center", "")
        row["Y_from_line"] = max_entry.get("Y_from_line", "")
        row["height"] = max_entry.get("height", "")
        row["focals_diff"] = max_entry.get("focals_diff", "")
        row["cam_angle_diff"] = max_entry.get("cam_angle_diff", "")
        row["can_improve_by_zoom"] = compute_can_improve_by_zoom(
            row["left_spare"], row["right_spare"], row["max_camera_overlap"]
        )
        row["decision"] = compute_decision(row["setup_severity"], row["can_improve_by_zoom"])
        row["setup_join_image"] = setup_joined_image
        rows.append(row)

    # Write output
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active

    all_headers = headers + new_columns
    ws.append(all_headers)

    setup_join_col = all_headers.index("setup_join_image") + 1  # 1-based column
    for row_idx, row in enumerate(rows, start=2):  # row 1 is header
        ws.append([row.get(h, "") for h in all_headers])
        url = row.get("setup_join_image", "")
        cell = ws.cell(row=row_idx, column=setup_join_col)
        if url and isinstance(url, str) and url.startswith("http"):
            cell.hyperlink = url
            cell.style = "Hyperlink"
        else:
            # Try local setup_join.jpg
            venue_id = str(row.get("PIXELLOT_VENUE_ID", "")).strip()
            if venue_id:
                setup_json_path = find_venue_setup_json(dataset_dir, venue_id)
                if setup_json_path:
                    jpg_path = setup_json_path.parent / "setup_join.jpg"
                    if jpg_path.is_file():
                        cell.hyperlink = jpg_path.resolve().as_uri()
                        cell.value = jpg_path.name
                        cell.font = Font(color="0000FF", underline="single")

    wb.save(output_path)
    matched = matched_max + matched_fallback
    unmatched = no_venue_dir + no_setup_json + no_setups + empty_venue_id
    print(f"\nPQS matched: {matched} (MAX: {matched_max}, fallback largest field_area: {matched_fallback})")
    print(f"PQS unmatched: {unmatched}")
    print(f"  Venue not in dataset: {no_venue_dir}")
    print(f"  No setup.json found: {no_setup_json}")
    print(f"  Empty setups array: {no_setups}")
    if empty_venue_id:
        print(f"  Empty venue ID: {empty_venue_id}")
    print(f"Dataset-only venues (not in PQS): {dataset_only}")
    if dataset_only_no_setups:
        print(f"  Dataset-only with empty setups array: {dataset_only_no_setups}")
    print(f"Total output rows: {len(rows)}")
    print(f"Output: {output_path}")


if __name__ == "__main__":
    main()

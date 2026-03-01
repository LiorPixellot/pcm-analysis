#!/usr/bin/env python3
"""Fix measurability annotations in complete_annotation.xlsx based on FN/FP review."""

from openpyxl import load_workbook

XLSX_PATH = "complete_annotation.xlsx"

# FN: GT was "yes" but should be "no" — keyed by (venue_id, checkup_date)
FN_KEYS = {
    ("5bed85304d2a230799c9373f", "2026-01-13 00:00:00"),
    ("5f69ec65df4978686c08e35e", "2026-02-09 00:00:00"),
    ("6061904edba9fc7bff81377d", "2026-02-09 00:00:00"),
    ("606989e9dba9fce376816950", "2026-02-09 00:00:00"),
    ("60698a05f059a9a1751c95bc", "2026-02-09 00:00:00"),
}

# FP: GT was "no" but should be "yes" — keyed by (venue_id, checkup_date)
FP_KEYS = {
    ("5b0d0ebe1bdc76b10759261e", "2026-02-04 00:00:00"),
    ("5b7a6888d01add079ed91403", "2026-02-04 00:00:00"),
    ("5baa31552b4397078488d332", "2026-02-04 00:00:00"),
    ("5bfabaa8b9f3e71e544e44b8", "2026-01-13 00:00:00"),
    ("5bfac2507a75205380d5b04b", "2026-01-13 00:00:00"),
    ("5c0cd1ba9d429307ab4f9ce1", "2026-01-13 00:00:00"),
    ("5c16745c42d75607bca8bf59", "2026-01-13 00:00:00"),
    ("5c1a191242d75607bca8cb41", "2026-01-13 00:00:00"),
    ("5c924e21745dcf4242809815", "2026-01-13 00:00:00"),
    ("5cbc4a8f7b7f4107a141d632", "2026-01-13 00:00:00"),
    ("5cbc57ba118884077f6cc80e", "2026-01-13 00:00:00"),
    ("5cbc5fd97b7f4107a141d655", "2026-01-13 00:00:00"),
    ("5cbd838b7b7f4107a141d735", "2026-01-13 00:00:00"),
    ("5cbdaca67b7f4107a141d76a", "2026-01-13 00:00:00"),
    ("5f686bdcde474e850bdf9b25", "2026-02-09 00:00:00"),
    ("5f6b5c147c47cab0e2e98f3d", "2026-02-09 00:00:00"),
    ("5f6c47ff656ea99e458cf6c6", "2026-02-09 00:00:00"),
    ("6061837db7465546da7626b6", "2026-02-09 00:00:00"),
    ("6062e082d5c26807244c1126", "2026-02-09 00:00:00"),
    ("6062e5df00a36fe4d5885226", "2026-02-09 00:00:00"),
    ("606411c50d2eb3219606a520", "2026-02-09 00:00:00"),
    ("606949450d2eb3722906c2bf", "2026-02-09 00:00:00"),
    ("606968e0f059a93dcd1c959a", "2026-02-09 00:00:00"),
    ("6069a7646d94cb24abb57f1b", "2026-02-09 00:00:00"),
}


def main():
    wb = load_workbook(XLSX_PATH)
    ws = wb.active

    # Find column indices from header
    header = {cell.value: cell.column for cell in ws[1]}
    sys_id_col = header["System ID"]
    date_col = header["Checkup date"]
    meas_col = header["Is Measurable Focus Image (Yes/No)"]

    fn_changed = []
    fp_changed = []

    for row in ws.iter_rows(min_row=2):
        venue_id = str(row[sys_id_col - 1].value).strip()
        date_val = str(row[date_col - 1].value).strip()
        cell = row[meas_col - 1]
        old_val = str(cell.value).strip()
        key = (venue_id, date_val)

        if key in FN_KEYS:
            cell.value = "no"
            fn_changed.append((venue_id, date_val, old_val, "no"))
        elif key in FP_KEYS:
            cell.value = "yes"
            fp_changed.append((venue_id, date_val, old_val, "yes"))

    wb.save(XLSX_PATH)

    print(f"FN → changed to 'no' ({len(fn_changed)} rows):")
    for vid, date, old, new in fn_changed:
        print(f"  {vid}  {date}  '{old}' → '{new}'")

    print(f"\nFP → changed to 'yes' ({len(fp_changed)} rows):")
    for vid, date, old, new in fp_changed:
        print(f"  {vid}  {date}  '{old}' → '{new}'")

    print(f"\nTotal changes: {len(fn_changed) + len(fp_changed)}")


if __name__ == "__main__":
    main()

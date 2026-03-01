#!/usr/bin/env python3
"""Merge Linux setup xlsx with Windows blur+setup xlsx (full outer join by venue_id)."""

import argparse
import os
from pathlib import Path

import openpyxl
from openpyxl.styles import Font
import pandas as pd

from enrich_blur_with_setup import (
    find_venue_setup_json, extract_max_setup, compute_setup_severity,
    compute_can_improve_by_zoom, compute_decision,
)
from add_movement_to_blur import load_movement_data, classify_movement_severity


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("linux_xlsx", nargs="+", help="One or more Linux setup xlsx files")
    parser.add_argument("--windows-xlsx", required=True, help="Windows blur+movement xlsx file")
    parser.add_argument("--dataset", nargs="+", default=None,
                        help="Dataset directories to scan for missing venues (optional, multiple allowed)")
    parser.add_argument("-o", "--output", default="data/linux_windows/merged_setup_blur.xlsx")
    args = parser.parse_args()

    # --- Extract hyperlinks from Linux sources before pandas loses them ---
    venue_hyperlinks = {}  # venue_id -> {col_name: target}
    for linux_file in args.linux_xlsx:
        linux_wb = openpyxl.load_workbook(linux_file)
        linux_ws = linux_wb.active
        linux_header = {cell.value: cell.column for cell in linux_ws[1]}
        venue_col = linux_header.get("venue_id") or linux_header.get("PIXELLOT_VENUE_ID")
        hyperlink_cols = {}
        for col_name, col_idx in linux_header.items():
            for row in range(2, linux_ws.max_row + 1):
                cell = linux_ws.cell(row=row, column=col_idx)
                if cell.hyperlink:
                    hyperlink_cols[col_name] = col_idx
                    break
        print(f"Hyperlink columns in {Path(linux_file).name}: {list(hyperlink_cols.keys())}")
        if hyperlink_cols:
            for row in range(2, linux_ws.max_row + 1):
                vid = linux_ws.cell(row=row, column=venue_col).value
                if vid is None:
                    continue
                for col_name, col_idx in hyperlink_cols.items():
                    cell = linux_ws.cell(row=row, column=col_idx)
                    if cell.hyperlink:
                        target = cell.hyperlink.target
                        venue_hyperlinks.setdefault(vid, {})[col_name] = target
        linux_wb.close()

    # --- Pandas merge ---
    linux_dfs = []
    for linux_file in args.linux_xlsx:
        df = pd.read_excel(linux_file)
        df.rename(columns={"PIXELLOT_VENUE_ID": "venue_id", "ACTIVE_CALIBRATIONS": "active_calibrations"}, inplace=True)
        print(f"Linux {Path(linux_file).name}: {len(df)} rows, {len(df.columns)} columns")
        linux_dfs.append(df)
    linux = pd.concat(linux_dfs, ignore_index=True)
    # Drop duplicate venue_ids — keep first occurrence (first file takes priority)
    linux = linux.drop_duplicates(subset="venue_id", keep="first")
    print(f"Linux combined: {len(linux)} rows (after dedup)")

    windows = pd.read_excel(args.windows_xlsx)
    print(f"Windows: {len(windows)} rows, {len(windows.columns)} columns")

    # Normalize key columns
    windows.rename(columns={"PIXELLOT_VENUE_ID": "venue_id", "ACTIVE_CALIBRATIONS": "active_calibrations"}, inplace=True)

    # Identify shared columns (besides venue_id)
    shared = [c for c in linux.columns if c in windows.columns and c != "venue_id"]
    print(f"Shared columns ({len(shared)}): {shared}")

    merged = windows.merge(linux, on="venue_id", how="outer", suffixes=("_windows", "_linux"))

    # For shared columns: prefer Linux, fallback to Windows
    for col in shared:
        merged[col] = merged[f"{col}_linux"].combine_first(merged[f"{col}_windows"])
        merged.drop(columns=[f"{col}_linux", f"{col}_windows"], inplace=True)

    Path(args.output).parent.mkdir(parents=True, exist_ok=True)
    merged.to_excel(args.output, index=False)

    # --- Re-apply hyperlinks with openpyxl ---
    wb = openpyxl.load_workbook(args.output)
    ws = wb.active
    out_header = {cell.value: cell.column for cell in ws[1]}
    vid_col = out_header["venue_id"]
    link_font = Font(color="0000FF", underline="single")
    links_applied = 0

    for row in range(2, ws.max_row + 1):
        vid = ws.cell(row=row, column=vid_col).value

        # Re-apply Linux file:// hyperlinks
        links = venue_hyperlinks.get(vid)
        if links:
            for col_name, target in links.items():
                col_idx = out_header.get(col_name)
                if col_idx is None:
                    continue
                cell = ws.cell(row=row, column=col_idx)
                cell.hyperlink = target
                cell.font = link_font
                links_applied += 1
                continue

        # Make any plain-text http(s) URLs clickable
        for col_name, col_idx in out_header.items():
            cell = ws.cell(row=row, column=col_idx)
            val = str(cell.value) if cell.value else ""
            if val.startswith(("http://", "https://")) and not cell.hyperlink:
                cell.hyperlink = val
                cell.font = link_font
                links_applied += 1

    # --- Second pass: enrich existing rows missing setup + add dataset-only venues ---
    dataset_only = 0
    setup_enriched = 0
    if args.dataset:
        setup_cols = {
            "setup_sport_type", "max_camera_overlap", "all_spares", "setup_severity",
            "left_spare", "right_spare", "s2_mode_calc", "X_of_center", "Y_from_line",
            "height", "focals_diff", "cam_angle_diff", "can_improve_by_zoom", "decision",
            "setup_join_image",
        }
        setup_join_col_idx = out_header.get("setup_join_image")
        setup_sev_col = out_header.get("setup_severity")

        def _extract_setup(dataset_dir, venue_id):
            """Extract setup data dict from dataset for a venue."""
            setup_json_path = find_venue_setup_json(dataset_dir, venue_id)
            setup_row = {}
            if not setup_json_path:
                return setup_row, None
            max_entry, used_fallback, setup_joined_image = extract_max_setup(setup_json_path)
            if max_entry:
                setup_row["setup_sport_type"] = max_entry.get("sport_type", "")
                setup_row["max_camera_overlap"] = max_entry.get("max_camera_overlap", "")
                setup_row["all_spares"] = max_entry.get("all_spares", "")
                setup_row["setup_severity"] = compute_setup_severity(setup_row["all_spares"])
                setup_row["left_spare"] = max_entry.get("left_spare", "")
                setup_row["right_spare"] = max_entry.get("right_spare", "")
                setup_row["s2_mode_calc"] = max_entry.get("s2_mode_calc", "")
                setup_row["X_of_center"] = max_entry.get("X_of_center", "")
                setup_row["Y_from_line"] = max_entry.get("Y_from_line", "")
                setup_row["height"] = max_entry.get("height", "")
                setup_row["focals_diff"] = max_entry.get("focals_diff", "")
                setup_row["cam_angle_diff"] = max_entry.get("cam_angle_diff", "")
                setup_row["can_improve_by_zoom"] = compute_can_improve_by_zoom(
                    setup_row["left_spare"], setup_row["right_spare"],
                    setup_row["max_camera_overlap"]
                )
                setup_row["decision"] = compute_decision(
                    setup_row["setup_severity"], setup_row["can_improve_by_zoom"]
                )
                setup_row["setup_join_image"] = setup_joined_image
            else:
                setup_row["setup_join_image"] = setup_joined_image
            return setup_row, setup_json_path

        def _apply_setup_hyperlink(row_num, setup_row, setup_json_path):
            """Apply hyperlink to setup_join_image cell."""
            if not setup_join_col_idx or not setup_row.get("setup_join_image"):
                return 0
            url = setup_row["setup_join_image"]
            cell = ws.cell(row=row_num, column=setup_join_col_idx)
            if isinstance(url, str) and url.startswith("http"):
                cell.hyperlink = url
                cell.font = link_font
                return 1
            # Try local setup_join.jpg
            if setup_json_path:
                jpg_path = setup_json_path.parent / "setup_join.jpg"
                if jpg_path.is_file():
                    cell.hyperlink = jpg_path.resolve().as_uri()
                    cell.value = jpg_path.name
                    cell.font = link_font
                    return 1
            return 0

        for dataset_path in args.dataset:
            dataset_dir = Path(dataset_path)
            if not dataset_dir.is_dir():
                print(f"Warning: dataset directory {dataset_dir} not found, skipping")
                continue

            print(f"\nProcessing dataset: {dataset_dir}")

            # Collect venue IDs already in the merged output, and rows missing setup data
            existing_vids = set()
            vids_missing_setup = {}  # venue_id -> list of row numbers
            for row in range(2, ws.max_row + 1):
                v = ws.cell(row=row, column=vid_col).value
                if v is None:
                    continue
                vid_str = str(v).strip()
                existing_vids.add(vid_str)
                if setup_sev_col:
                    sev = ws.cell(row=row, column=setup_sev_col).value
                    if not sev or not str(sev).strip():
                        vids_missing_setup.setdefault(vid_str, []).append(row)

            # Load movement data from dataset
            movement_data = load_movement_data(dataset_dir)

            # Part 1: Enrich existing rows that are missing setup data
            ds_enriched = 0
            for venue_id, row_nums in vids_missing_setup.items():
                setup_row, setup_json_path = _extract_setup(dataset_dir, venue_id)
                if not setup_row or "setup_severity" not in setup_row:
                    continue
                ds_enriched += 1
                for row_num in row_nums:
                    for col_name, col_idx in out_header.items():
                        if col_name in setup_cols and col_name in setup_row:
                            ws.cell(row=row_num, column=col_idx, value=setup_row[col_name])
                    links_applied += _apply_setup_hyperlink(row_num, setup_row, setup_json_path)
            setup_enriched += ds_enriched

            # Part 2: Add new rows for venues not in merged output at all
            ds_added = 0
            next_row = ws.max_row + 1
            for venue_dir_entry in sorted(os.listdir(dataset_dir)):
                venue_path = dataset_dir / venue_dir_entry
                if not venue_path.is_dir():
                    continue
                venue_id = venue_dir_entry
                if venue_id in existing_vids:
                    continue

                setup_row, setup_json_path = _extract_setup(dataset_dir, venue_id)
                venue_movements = movement_data.get(venue_id)
                calibrations = list(venue_movements.items()) if venue_movements else [(None, None)]

                for calib, mov in calibrations:
                    ws.cell(row=next_row, column=vid_col, value=venue_id)
                    # Write setup columns
                    for col_name, col_idx in out_header.items():
                        if col_name in setup_cols and col_name in setup_row:
                            ws.cell(row=next_row, column=col_idx, value=setup_row[col_name])
                    # Write movement columns
                    if mov:
                        for col_name, col_idx in out_header.items():
                            if col_name == "calibration_movement":
                                ws.cell(row=next_row, column=col_idx, value=calib)
                            elif col_name == "movement_indicator":
                                ws.cell(row=next_row, column=col_idx, value=mov["movement_indicator"])
                            elif col_name == "movement_length_cam0":
                                ws.cell(row=next_row, column=col_idx, value=mov["movement_length_cam0"])
                            elif col_name == "movement_length_cam1":
                                ws.cell(row=next_row, column=col_idx, value=mov["movement_length_cam1"])
                            elif col_name == "movement_length_cam2":
                                ws.cell(row=next_row, column=col_idx, value=mov["movement_length_cam2"])
                            elif col_name == "movement_severity":
                                ws.cell(row=next_row, column=col_idx, value=classify_movement_severity(
                                    mov["movement_indicator"], mov["movement_length_cam0"],
                                    mov["movement_length_cam1"], mov["movement_length_cam2"]
                                ))
                    links_applied += _apply_setup_hyperlink(next_row, setup_row, setup_json_path)
                    next_row += 1

                ds_added += 1
                existing_vids.add(venue_id)  # prevent duplicates across datasets

            dataset_only += ds_added
            print(f"  Enriched existing rows: {ds_enriched} venues")
            print(f"  Added dataset-only venues: {ds_added}")

        print(f"\nTotal enriched: {setup_enriched} venues, total added: {dataset_only} venues")

    wb.save(args.output)
    print(f"Hyperlinks applied: {links_applied}")

    total_rows = ws.max_row - 1  # exclude header
    print(f"Merged: {total_rows} rows, {len(merged.columns)} columns → {args.output}")


if __name__ == "__main__":
    main()
